from __future__ import annotations

import os
import shutil
import subprocess
import sys
import traceback
import json
import zipfile
import tempfile
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image as PILImage, ImageTk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether, Image as RLImage
)


APP_NAME = "Sistema de Gestão do Claviculário - ECL"
APP_VERSION = "0.9"
BASE_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
BACKUP_DIR = BASE_DIR / "backups"
PROTOCOL_DIR = BASE_DIR / "protocolos"
REPORT_DIR = BASE_DIR / "relatorios"
LOG_DIR = BASE_DIR / "logs"
ASSET_DIR = BASE_DIR / "assets"
LOGO_HEADER_PATH = ASSET_DIR / "logo_header.png"
LOGO_PDF_PATH = ASSET_DIR / "logo_pdf.png"
DB_FILE = DATA_DIR / "Claviculario_ECL.xlsx"
CONFIG_FILE = BASE_DIR / "config.json"

CATEGORIAS = [
    "Administrativo",
    "Módulo Laranja",
    "Módulo Azul",
    "Módulo Amarelo",
    "Operacional",
]

MOTIVOS_INATIVACAO = [
    "Loja encerrada",
    "Demolição",
    "Mudança de operação",
    "Chave substituída",
    "Outro",
]


def now_text() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def log_error(exc: BaseException) -> None:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / "erros.log"
    with log_file.open("a", encoding="utf-8") as f:
        f.write("\n" + "=" * 90 + "\n")
        f.write(now_text() + "\n")
        f.write("".join(traceback.format_exception(exc)))


def open_file(path: Path) -> None:
    """Abre arquivo ou pasta usando o aplicativo padrão do sistema."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


def load_app_config() -> dict:
    if not CONFIG_FILE.exists():
        return {}
    try:
        return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_app_config(config: dict) -> None:
    CONFIG_FILE.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def google_backup_dir() -> Path | None:
    cfg = load_app_config()
    raw = str(cfg.get("google_drive_folder") or "").strip()
    if not raw:
        return None
    return Path(raw) / "Claviculario ECL" / "Backups"


def create_full_backup_zip(destination: Path | None = None) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    local_path = BACKUP_DIR / f"Claviculario_ECL_Backup_{stamp}.zip"

    with zipfile.ZipFile(local_path, "w", zipfile.ZIP_DEFLATED) as z:
        if DB_FILE.exists():
            z.write(DB_FILE, "data/Claviculario_ECL.xlsx")

        for folder, arc_folder in [
            (PROTOCOL_DIR, "protocolos"),
            (REPORT_DIR, "relatorios"),
        ]:
            if folder.exists():
                for item in folder.rglob("*"):
                    if item.is_file():
                        z.write(item, f"{arc_folder}/{item.relative_to(folder)}")

    if destination:
        destination.mkdir(parents=True, exist_ok=True)
        drive_copy = destination / local_path.name
        shutil.copy2(local_path, drive_copy)
        return drive_copy
    return local_path


def restore_backup_zip(zip_path: Path) -> None:
    zip_path = Path(zip_path)
    if not zip_path.exists():
        raise FileNotFoundError("Backup não encontrado.")

    # Backup de segurança antes de restaurar.
    if DB_FILE.exists():
        create_full_backup_zip()

    with tempfile.TemporaryDirectory(prefix="claviculario_restore_") as tmp:
        tmp_dir = Path(tmp)
        with zipfile.ZipFile(zip_path, "r") as z:
            z.extractall(tmp_dir)

        restored_db = tmp_dir / "data" / "Claviculario_ECL.xlsx"
        if not restored_db.exists():
            raise ValueError("O arquivo selecionado não contém uma base válida do Claviculário ECL.")

        DATA_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copy2(restored_db, DB_FILE)

        for folder_name, target in [
            ("protocolos", PROTOCOL_DIR),
            ("relatorios", REPORT_DIR),
        ]:
            source = tmp_dir / folder_name
            if source.exists():
                target.mkdir(parents=True, exist_ok=True)
                for item in source.rglob("*"):
                    if item.is_file():
                        out = target / item.relative_to(source)
                        out.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(item, out)


class MapTooltip:
    def __init__(self, widget, text_provider):
        self.widget = widget
        self.text_provider = text_provider
        self.tip = None
        widget.bind("<Enter>", self.show, add="+")
        widget.bind("<Leave>", self.hide, add="+")

    def show(self, event=None):
        self.hide()
        text = self.text_provider()
        if not text:
            return
        x = self.widget.winfo_rootx() + self.widget.winfo_width() + 8
        y = self.widget.winfo_rooty()
        self.tip = tk.Toplevel(self.widget)
        self.tip.wm_overrideredirect(True)
        self.tip.wm_geometry(f"+{x}+{y}")
        label = tk.Label(
            self.tip, text=text, justify="left",
            bg="#17324D", fg="white",
            relief="solid", borderwidth=1,
            font=("Segoe UI", 9), padx=10, pady=8
        )
        label.pack()

    def hide(self, event=None):
        if self.tip is not None:
            self.tip.destroy()
            self.tip = None


class ExcelRepository:
    SHEETS = {
        "CHAVES_ECL": [
            "Número no Claviculário", "Código", "Descrição", "Categoria/Módulo", "Local", "Cópias",
            "Status Operacional", "Ativa", "Observação",
            "Criado em", "Atualizado em", "Motivo inativação"
        ],
        "MOVIMENTACOES": [
            "Protocolo", "Número no Claviculário", "Código ECL", "Descrição", "Categoria/Módulo",
            "Responsável", "Setor", "Motivo", "Observação",
            "Data/Hora Retirada", "Data/Hora Devolução", "Status"
        ],
        "PROTOCOLOS": [
            "Protocolo", "Número no Claviculário", "Código ECL", "Descrição", "Categoria/Módulo",
            "Responsável", "Setor", "Motivo", "Observação",
            "Data/Hora Retirada", "Data/Hora Devolução", "Status", "Arquivo PDF"
        ],
        "CHAVES_INATIVAS": [
            "Número no Claviculário", "Código ECL", "Descrição", "Categoria/Módulo", "Local",
            "Motivo", "Data inativação"
        ],
        "CONFIGURACOES": ["Chave", "Valor"],
    }

    def __init__(self, path: Path):
        self.path = path
        for d in [DATA_DIR, BACKUP_DIR, PROTOCOL_DIR, REPORT_DIR, LOG_DIR, ASSET_DIR]:
            d.mkdir(parents=True, exist_ok=True)
        self.ensure_database()

    def ensure_database(self) -> None:
        if self.path.exists():
            self._upgrade_database()
            return

        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, headers in self.SHEETS.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            self._style_header(ws)
            ws.freeze_panes = "A2"

        wb["CONFIGURACOES"].append(["ultimo_protocolo", "0"])
        self._auto_width(wb)
        self._save_atomic(wb)

    def _upgrade_database(self) -> None:
        """Atualiza abas/colunas preservando os dados das versões anteriores."""
        wb = self.load_raw()
        changed = False

        for sheet_name, headers in self.SHEETS.items():
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
                self._style_header(ws)
                ws.freeze_panes = "A2"
                changed = True
                continue

            ws = wb[sheet_name]
            old_headers = [str(c.value or "").strip() for c in ws[1]]

            # Quando a ordem/quantidade de colunas muda, reconstrói a tabela
            # usando o nome dos campos, para não deslocar dados antigos.
            if old_headers != headers:
                old_rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        old_rows.append(dict(zip(old_headers, row)))

                if ws.max_row:
                    ws.delete_rows(1, ws.max_row)

                ws.append(headers)
                for item in old_rows:
                    ws.append([item.get(header, "") for header in headers])

                changed = True

            self._style_header(ws)
            ws.freeze_panes = "A2"

        cfg = wb["CONFIGURACOES"]
        has_counter = any(
            str(row[0] or "").strip() == "ultimo_protocolo"
            for row in cfg.iter_rows(min_row=2, values_only=True)
        )
        if not has_counter:
            cfg.append(["ultimo_protocolo", "0"])
            changed = True

        if changed:
            self._auto_width(wb)
            self._save_atomic(wb)
        else:
            wb.close()

    @staticmethod
    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17324D")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, wb) -> None:
        for ws in wb.worksheets:
            for col in ws.columns:
                values = [len(str(c.value)) if c.value is not None else 0 for c in col[:200]]
                max_len = max(values + [0])
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 36)

    def backup(self) -> None:
        if self.path.exists():
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
            shutil.copy2(self.path, BACKUP_DIR / f"Claviculario_ECL_{stamp}.xlsx")

    def _save_atomic(self, wb) -> None:
        tmp = self.path.with_name(self.path.stem + "_tmp.xlsx")
        try:
            wb.save(tmp)
            wb.close()
            tmp.replace(self.path)
        finally:
            try:
                if tmp.exists():
                    tmp.unlink()
            except OSError:
                pass

    def load_raw(self):
        try:
            return load_workbook(self.path)
        except PermissionError as e:
            raise RuntimeError(
                "A planilha Claviculario_ECL.xlsx está aberta no Excel. "
                "Feche a planilha e tente novamente."
            ) from e

    def load(self):
        return self.load_raw()

    def save(self, wb) -> None:
        self.backup()
        self._auto_width(wb)
        self._save_atomic(wb)

    @staticmethod
    def _rows_as_dicts(ws):
        headers = [c.value for c in ws[1]]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if any(v is not None for v in row):
                yield idx, dict(zip(headers, row))

    def list_keys(self, include_inactive: bool = False):
        wb = self.load()
        ws = wb["CHAVES_ECL"]
        result = []
        for _, item in self._rows_as_dicts(ws):
            if include_inactive or str(item.get("Ativa", "SIM")).upper() == "SIM":
                result.append(item)
        wb.close()
        return result

    def _validate_clavicular_number(self, numero) -> int:
        try:
            value = int(str(numero).strip())
        except (TypeError, ValueError):
            raise ValueError("Informe um número do claviculário entre 1 e 200.")
        if not 1 <= value <= 200:
            raise ValueError("O número do claviculário deve estar entre 1 e 200.")
        return value

    def add_key(self, numero, codigo, descricao, categoria, local, copias, observacao):
        numero = self._validate_clavicular_number(numero)
        codigo = codigo.strip().upper()
        descricao = descricao.strip()
        categoria = categoria.strip()
        local = local.strip()
        observacao = observacao.strip()
        try:
            copias = int(str(copias or "0").strip())
        except ValueError as e:
            raise ValueError("Cópias deve ser um número inteiro igual ou maior que zero.") from e
        if copias < 0:
            raise ValueError("Cópias não pode ser negativo.")

        if not codigo or not descricao or not categoria:
            raise ValueError("Código, descrição e categoria/módulo são obrigatórios.")

        wb = self.load()
        ws = wb["CHAVES_ECL"]
        for _, item in self._rows_as_dicts(ws):
            if str(item.get("Código", "")).strip().upper() == codigo and str(item.get("Ativa", "SIM")).upper() == "SIM":
                wb.close()
                raise ValueError(f"Já existe uma Chave do ECL ativa com o código {codigo}.")
            existing_number = item.get("Número no Claviculário")
            if (
                str(item.get("Ativa", "SIM")).upper() == "SIM"
                and existing_number not in (None, "")
                and int(existing_number) == numero
            ):
                wb.close()
                raise ValueError(
                    f"O número {numero} já está sendo utilizado pela chave "
                    f"{item.get('Código')} - {item.get('Descrição')}."
                )

        dt = now_text()
        ws.append([numero, codigo, descricao, categoria, local, copias, "Disponível", "SIM",
                   observacao, dt, dt, ""])
        self.save(wb)

    def update_key(self, original_code, numero, codigo, descricao, categoria, local, copias, observacao):
        numero = self._validate_clavicular_number(numero)
        codigo = codigo.strip().upper()
        descricao = descricao.strip()
        try:
            copias = int(str(copias or "0").strip())
        except ValueError as e:
            raise ValueError("Cópias deve ser um número inteiro igual ou maior que zero.") from e
        if copias < 0:
            raise ValueError("Cópias não pode ser negativo.")
        if not codigo or not descricao:
            raise ValueError("Código e descrição são obrigatórios.")

        wb = self.load()
        ws = wb["CHAVES_ECL"]
        headers = [c.value for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}
        target = None
        for idx, item in self._rows_as_dicts(ws):
            existing = str(item.get("Código", "")).strip().upper()
            if existing == original_code.strip().upper() and str(item.get("Ativa", "SIM")).upper() == "SIM":
                target = idx
            elif existing == codigo and str(item.get("Ativa", "SIM")).upper() == "SIM":
                wb.close()
                raise ValueError(f"Já existe uma Chave do ECL ativa com o código {codigo}.")

            existing_number = item.get("Número no Claviculário")
            if (
                existing != original_code.strip().upper()
                and str(item.get("Ativa", "SIM")).upper() == "SIM"
                and existing_number not in (None, "")
                and int(existing_number) == numero
            ):
                wb.close()
                raise ValueError(
                    f"O número {numero} já está sendo utilizado pela chave "
                    f"{item.get('Código')} - {item.get('Descrição')}."
                )

        if target is None:
            wb.close()
            raise ValueError("Chave do ECL não encontrada.")

        # Não deixa reduzir o estoque abaixo das unidades que estão fora.
        open_count = 0
        for _, mov in self._rows_as_dicts(wb["MOVIMENTACOES"]):
            if str(mov.get("Código ECL", "")).strip().upper() == original_code.strip().upper() and mov.get("Status") == "Retirada":
                open_count += 1
        total = 1 + copias
        if total < open_count:
            wb.close()
            raise ValueError(
                f"Existem {open_count} unidade(s) retiradas. O total físico não pode ser menor que isso."
            )

        ws.cell(target, col["Número no Claviculário"]).value = numero
        ws.cell(target, col["Código"]).value = codigo
        ws.cell(target, col["Descrição"]).value = descricao
        ws.cell(target, col["Categoria/Módulo"]).value = categoria
        ws.cell(target, col["Local"]).value = local.strip()
        ws.cell(target, col["Cópias"]).value = copias
        ws.cell(target, col["Observação"]).value = observacao.strip()
        ws.cell(target, col["Atualizado em"]).value = now_text()
        self._sync_key_status(wb, target, codigo)
        self.save(wb)

    def key_stock_info(self, codigo, wb=None):
        own_wb = wb is None
        if own_wb:
            wb = self.load()
        key = None
        for _, item in self._rows_as_dicts(wb["CHAVES_ECL"]):
            if str(item.get("Código", "")).strip().upper() == str(codigo).strip().upper() and str(item.get("Ativa", "SIM")).upper() == "SIM":
                key = item
                break
        if not key:
            if own_wb:
                wb.close()
            return None
        try:
            copies = int(key.get("Cópias") or 0)
        except (TypeError, ValueError):
            copies = 0
        total = 1 + max(copies, 0)
        withdrawn = 0
        for _, mov in self._rows_as_dicts(wb["MOVIMENTACOES"]):
            if str(mov.get("Código ECL", "")).strip().upper() == str(codigo).strip().upper() and mov.get("Status") == "Retirada":
                withdrawn += 1
        available = max(total - withdrawn, 0)
        if available == 0:
            status = "Sem chave disponível"
        elif withdrawn > 0:
            status = "Parcialmente retirada"
        else:
            status = "Disponível"
        result = {
            "copias": copies,
            "total": total,
            "retiradas": withdrawn,
            "disponiveis": available,
            "status": status,
        }
        if own_wb:
            wb.close()
        return result

    def _sync_key_status(self, wb, key_row, codigo):
        info = self.key_stock_info(codigo, wb=wb)
        if not info:
            return
        ws = wb["CHAVES_ECL"]
        headers = [c.value for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}
        ws.cell(key_row, col["Status Operacional"]).value = info["status"]
        ws.cell(key_row, col["Atualizado em"]).value = now_text()

    def free_key(self, codigo):
        """Libera a posição para reutilização e preserva todo o histórico."""
        wb = self.load()
        ws = wb["CHAVES_ECL"]
        headers = [c.value for c in ws[1]]
        col = {h: i + 1 for i, h in enumerate(headers)}
        target = None
        key = None
        for idx, item in self._rows_as_dicts(ws):
            if str(item.get("Código", "")).strip().upper() == str(codigo).strip().upper() and str(item.get("Ativa", "SIM")).upper() == "SIM":
                target, key = idx, item
                break
        if target is None:
            wb.close()
            raise ValueError("Chave ativa não encontrada.")

        dt = now_text()
        # Encerra retiradas em aberto como entrega definitiva/liberação.
        for idx, mov in self._rows_as_dicts(wb["MOVIMENTACOES"]):
            if str(mov.get("Código ECL", "")).strip().upper() == str(codigo).strip().upper() and mov.get("Status") == "Retirada":
                mov_headers = [c.value for c in wb["MOVIMENTACOES"][1]]
                mc = {h: i + 1 for i, h in enumerate(mov_headers)}
                wb["MOVIMENTACOES"].cell(idx, mc["Data/Hora Devolução"]).value = dt
                wb["MOVIMENTACOES"].cell(idx, mc["Status"]).value = "Entrega definitiva / posição liberada"
                protocol = mov.get("Protocolo")
                for pidx, prot in self._rows_as_dicts(wb["PROTOCOLOS"]):
                    if str(prot.get("Protocolo", "")) == str(protocol):
                        ph = [c.value for c in wb["PROTOCOLOS"][1]]
                        pc = {h: i + 1 for i, h in enumerate(ph)}
                        wb["PROTOCOLOS"].cell(pidx, pc["Data/Hora Devolução"]).value = dt
                        wb["PROTOCOLOS"].cell(pidx, pc["Status"]).value = "Entrega definitiva / posição liberada"
                        break

        ws.cell(target, col["Ativa"]).value = "NÃO"
        ws.cell(target, col["Status Operacional"]).value = "Livre"
        ws.cell(target, col["Motivo inativação"]).value = "Posição liberada no claviculário"
        ws.cell(target, col["Atualizado em"]).value = dt
        if "CHAVES_INATIVAS" in wb.sheetnames:
            wb["CHAVES_INATIVAS"].append([
                key.get("Número no Claviculário"), key.get("Código"), key.get("Descrição"),
                key.get("Categoria/Módulo"), key.get("Local"), "Posição liberada", dt
            ])
        self.save(wb)
        return key

    def inactivate_key(self, codigo, motivo):
        # Compatibilidade com versões antigas: agora equivale a liberar posição.
        return self.free_key(codigo)

    def restore_key(self, codigo):
        raise ValueError("A versão atual usa Espaço Livre. Cadastre novamente a chave na posição desejada.")

    def get_next_protocol(self, wb) -> str:
        cfg = wb["CONFIGURACOES"]
        row_idx = None
        current = 0
        for idx, item in self._rows_as_dicts(cfg):
            if str(item.get("Chave", "")).strip() == "ultimo_protocolo":
                row_idx = idx
                try:
                    current = int(item.get("Valor") or 0)
                except (TypeError, ValueError):
                    current = 0
                break

        if row_idx is None:
            cfg.append(["ultimo_protocolo", "0"])
            row_idx = cfg.max_row

        current += 1
        cfg.cell(row_idx, 2).value = str(current)
        return f"{datetime.now().year}-{current:06d}"

    def withdraw_key(self, codigo, responsavel, setor, motivo, observacao):
        responsavel = responsavel.strip()
        setor = setor.strip()
        motivo = motivo.strip()
        observacao = observacao.strip()
        if not responsavel:
            raise ValueError("Informe o responsável pela retirada.")
        if not setor:
            raise ValueError("Informe o setor.")
        if not motivo:
            raise ValueError("Informe o motivo da retirada.")

        wb = self.load()
        ws_keys = wb["CHAVES_ECL"]
        key_row = None
        key = None
        for idx, item in self._rows_as_dicts(ws_keys):
            if str(item.get("Código", "")).strip().upper() == codigo.strip().upper() and str(item.get("Ativa", "SIM")).upper() == "SIM":
                key_row, key = idx, item
                break
        if key_row is None or key is None:
            wb.close()
            raise ValueError("Chave do ECL não encontrada ou posição está livre.")

        stock = self.key_stock_info(codigo, wb=wb)
        if not stock or stock["disponiveis"] <= 0:
            wb.close()
            raise ValueError("Sem chave/cópia disponível para retirada.")

        protocol = self.get_next_protocol(wb)
        retirada = now_text()
        wb["MOVIMENTACOES"].append([
            protocol, key.get("Número no Claviculário"), key.get("Código"), key.get("Descrição"),
            key.get("Categoria/Módulo"), responsavel, setor, motivo, observacao,
            retirada, "", "Retirada"
        ])
        wb["PROTOCOLOS"].append([
            protocol, key.get("Número no Claviculário"), key.get("Código"), key.get("Descrição"),
            key.get("Categoria/Módulo"), responsavel, setor, motivo, observacao,
            retirada, "", "Retirada", ""
        ])
        self._sync_key_status(wb, key_row, codigo)
        self.save(wb)

        remaining = stock["disponiveis"] - 1
        return {
            "protocolo": protocol,
            "numero": key.get("Número no Claviculário"),
            "codigo": key.get("Código"),
            "descricao": key.get("Descrição"),
            "categoria": key.get("Categoria/Módulo"),
            "local": key.get("Local"),
            "copias": stock["copias"],
            "total": stock["total"],
            "disponiveis_apos": remaining,
            "responsavel": responsavel,
            "setor": setor,
            "motivo": motivo,
            "observacao": observacao,
            "retirada": retirada,
        }

    def set_protocol_pdf(self, protocol: str, pdf_path: Path):
        wb = self.load()
        ws = wb["PROTOCOLOS"]
        for idx, item in self._rows_as_dicts(ws):
            if str(item.get("Protocolo", "")) == protocol:
                # Arquivo PDF é a coluna 12 no schema atual
                ws.cell(idx, 13).value = str(pdf_path)
                self.save(wb)
                return
        wb.close()

    def list_open_withdrawals(self):
        wb = self.load()
        ws = wb["MOVIMENTACOES"]
        result = []
        for _, item in self._rows_as_dicts(ws):
            if str(item.get("Status", "")) == "Retirada":
                result.append(item)
        wb.close()
        return result

    def return_key(self, protocol: str):
        wb = self.load()
        ws_mov = wb["MOVIMENTACOES"]
        movement_row = None
        movement = None
        for idx, item in self._rows_as_dicts(ws_mov):
            if str(item.get("Protocolo", "")) == protocol:
                movement_row, movement = idx, item
                break
        if movement_row is None or movement is None:
            wb.close()
            raise ValueError("Movimentação não encontrada.")
        if movement.get("Status") != "Retirada":
            wb.close()
            raise ValueError("Esta movimentação já foi encerrada.")

        dt = now_text()
        mh = [c.value for c in ws_mov[1]]
        mc = {h: i + 1 for i, h in enumerate(mh)}
        ws_mov.cell(movement_row, mc["Data/Hora Devolução"]).value = dt
        ws_mov.cell(movement_row, mc["Status"]).value = "Devolvida"

        code = str(movement.get("Código ECL", ""))
        ws_keys = wb["CHAVES_ECL"]
        for idx, item in self._rows_as_dicts(ws_keys):
            if str(item.get("Código", "")).strip().upper() == code.strip().upper() and str(item.get("Ativa", "SIM")).upper() == "SIM":
                self._sync_key_status(wb, idx, code)
                break

        ws_protocols = wb["PROTOCOLOS"]
        ph = [c.value for c in ws_protocols[1]]
        pc = {h: i + 1 for i, h in enumerate(ph)}
        for idx, item in self._rows_as_dicts(ws_protocols):
            if str(item.get("Protocolo", "")) == protocol:
                ws_protocols.cell(idx, pc["Data/Hora Devolução"]).value = dt
                ws_protocols.cell(idx, pc["Status"]).value = "Devolvida"
                break

        self.save(wb)
        movement["Data/Hora Devolução"] = dt
        movement["Status"] = "Devolvida"
        return movement

    def list_movements(self, query: str = ""):
        wb = self.load()
        ws = wb["MOVIMENTACOES"]
        result = [item for _, item in self._rows_as_dicts(ws)]
        wb.close()
        result.reverse()

        q = query.strip().lower()
        if q:
            filtered = []
            for item in result:
                text = " ".join(str(v or "") for v in item.values()).lower()
                if q in text:
                    filtered.append(item)
            result = filtered
        return result

    def list_protocols(self, query: str = ""):
        wb = self.load()
        ws = wb["PROTOCOLOS"]
        result = [item for _, item in self._rows_as_dicts(ws)]
        wb.close()
        result.reverse()
        q = query.strip().lower()
        if q:
            result = [
                item for item in result
                if q in " ".join(str(v or "") for v in item.values()).lower()
            ]
        return result

    def get_protocol(self, protocol: str):
        wb = self.load()
        ws = wb["PROTOCOLOS"]
        for _, item in self._rows_as_dicts(ws):
            if str(item.get("Protocolo", "")) == protocol:
                wb.close()
                return item
        wb.close()
        return None


    @staticmethod
    def _parse_datetime_text(value):
        if not value:
            return None
        try:
            return datetime.strptime(str(value), "%d/%m/%Y %H:%M:%S")
        except (TypeError, ValueError):
            return None

    def filtered_movements(self, start_date="", end_date="", categoria="", status="", search=""):
        movements = self.list_movements()
        start_dt = None
        end_dt = None

        if start_date.strip():
            try:
                start_dt = datetime.strptime(start_date.strip(), "%d/%m/%Y")
            except ValueError as e:
                raise ValueError("Data inicial inválida. Use DD/MM/AAAA.") from e

        if end_date.strip():
            try:
                end_dt = datetime.strptime(end_date.strip(), "%d/%m/%Y")
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
            except ValueError as e:
                raise ValueError("Data final inválida. Use DD/MM/AAAA.") from e

        categoria = categoria.strip()
        status = status.strip()
        search = search.strip().lower()

        result = []
        for item in movements:
            retirada_dt = self._parse_datetime_text(item.get("Data/Hora Retirada"))
            if start_dt and (not retirada_dt or retirada_dt < start_dt):
                continue
            if end_dt and (not retirada_dt or retirada_dt > end_dt):
                continue
            if categoria and categoria != "Todos" and item.get("Categoria/Módulo") != categoria:
                continue
            if status and status != "Todos" and item.get("Status") != status:
                continue
            if search:
                blob = " ".join(str(v or "") for v in item.values()).lower()
                if search not in blob:
                    continue
            result.append(item)
        return result

    def dashboard_counts(self):
        active = self.list_keys()
        total = len(active)
        with_available = 0
        with_withdrawals = 0
        for k in active:
            info = self.key_stock_info(k.get("Código"))
            if info and info["disponiveis"] > 0:
                with_available += 1
            if info and info["retiradas"] > 0:
                with_withdrawals += 1
        free_spaces = max(200 - total, 0)
        return total, with_available, with_withdrawals, free_spaces

    def category_counts(self):
        counts = {c: 0 for c in CATEGORIAS}
        for item in self.list_keys():
            cat = item.get("Categoria/Módulo")
            counts[cat] = counts.get(cat, 0) + 1
        return counts


def generate_protocol_pdf(data: dict) -> Path:
    PROTOCOL_DIR.mkdir(parents=True, exist_ok=True)
    safe_protocol = str(data["protocolo"]).replace("/", "-")
    path = PROTOCOL_DIR / f"Protocolo_{safe_protocol}_{data['codigo']}.pdf"

    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Protocolo {data['protocolo']}",
        author="Claviculário ECL",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ProtocolTitle", parent=styles["Title"], alignment=TA_CENTER,
        fontName="Helvetica-Bold", fontSize=15, leading=19, textColor=colors.HexColor("#17324D")
    )
    subtitle_style = ParagraphStyle(
        "ProtocolSubtitle", parent=styles["Heading2"], alignment=TA_CENTER,
        fontName="Helvetica-Bold", fontSize=12, leading=15
    )
    normal = ParagraphStyle(
        "ProtocolNormal", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=10, leading=14
    )
    small = ParagraphStyle(
        "ProtocolSmall", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=8.5, leading=11, textColor=colors.HexColor("#51606D")
    )

    story = []
    if LOGO_PDF_PATH.exists():
        logo = RLImage(str(LOGO_PDF_PATH), width=72 * mm, height=21 * mm)
        logo.hAlign = "CENTER"
        story.extend([logo, Spacer(1, 3 * mm)])
    story.extend([
        Paragraph("SISTEMA DE GESTÃO DO CLAVICULÁRIO - ECL", title_style),
        Spacer(1, 4 * mm),
        Paragraph("PROTOCOLO DE RETIRADA DE CHAVE DO ECL", subtitle_style),
        Paragraph(f"Nº {data['protocolo']}", subtitle_style),
        Spacer(1, 8 * mm),
    ])

    info = [
        ["Número no Claviculário", f"{int(data['numero']):03d}" if data.get("numero") not in (None, "") else ""],
        ["Chave do ECL", str(data["codigo"])],
        ["Descrição", str(data["descricao"])],
        ["Categoria/Módulo", str(data["categoria"])],
        ["Local", str(data.get("local") or "")],
        ["Cópias cadastradas", str(data.get("copias", 0)) if data.get("copias", 0) else "Sem cópia"],
        ["Unidades disponíveis após retirada", str(data.get("disponiveis_apos", ""))],
        ["Responsável", str(data["responsavel"])],
        ["Setor", str(data["setor"])],
        ["Motivo da retirada", str(data["motivo"])],
        ["Data/Hora da retirada", str(data["retirada"])],
        ["Observações", str(data.get("observacao") or "")],
    ]
    table = Table(info, colWidths=[48 * mm, 102 * mm], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E9EFF5")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#17324D")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#A9B5C0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    story.append(Spacer(1, 10 * mm))

    declaration = (
        "Declaro que recebi a Chave do ECL acima identificada e assumo a "
        "responsabilidade por sua guarda, conservação e posterior devolução."
    )
    story.append(Paragraph(declaration, normal))
    story.append(Spacer(1, 18 * mm))

    signature_data = [
        ["____________________________________", "____________________________________"],
        ["Assinatura de quem retirou", "Responsável pela entrega"],
        [str(data["responsavel"]), ""],
    ]
    sig = Table(signature_data, colWidths=[75 * mm, 75 * mm], hAlign="CENTER")
    sig.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(KeepTogether(sig))
    story.append(Spacer(1, 18 * mm))
    story.append(Paragraph(
        f"Documento gerado pelo {APP_NAME} - versão {APP_VERSION}.",
        small
    ))

    doc.build(story)
    return path


def generate_movement_report_pdf(rows, filters: dict) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = REPORT_DIR / f"Relatorio_Movimentacoes_{stamp}.pdf"

    doc = SimpleDocTemplate(
        str(path), pagesize=A4, rightMargin=10 * mm, leftMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Relatório de Movimentações - Claviculário ECL",
        author="Claviculário ECL",
    )

    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], alignment=TA_CENTER,
        fontName="Helvetica-Bold", fontSize=15, leading=18,
        textColor=colors.HexColor("#17324D")
    )
    small = ParagraphStyle(
        "ReportSmall", parent=styles["BodyText"], fontSize=8.5, leading=11
    )

    story = []
    if LOGO_PDF_PATH.exists():
        logo = RLImage(str(LOGO_PDF_PATH), width=72 * mm, height=21 * mm)
        logo.hAlign = "CENTER"
        story.extend([logo, Spacer(1, 3 * mm)])
    story.extend([
        Paragraph("RELATÓRIO DE MOVIMENTAÇÕES - CLAVICULÁRIO ECL", title),
        Spacer(1, 4 * mm),
        Paragraph(
            f"Período: {filters.get('inicio') or 'Todos'} até {filters.get('fim') or 'Todos'} | "
            f"Módulo: {filters.get('categoria') or 'Todos'} | "
            f"Status: {filters.get('status') or 'Todos'}",
            small
        ),
        Paragraph(f"Filtro de pesquisa: {filters.get('search') or 'Nenhum'}", small),
        Paragraph(f"Gerado em: {now_text()} | Total de registros: {len(rows)}", small),
        Spacer(1, 5 * mm),
    ])

    data = [["Protocolo", "Nº", "Chave", "Responsável", "Retirada", "Devolução", "Status"]]
    for item in rows:
        number = item.get("Número no Claviculário")
        number_txt = f"{int(number):03d}" if number not in (None, "") else ""
        data.append([
            str(item.get("Protocolo") or ""),
            number_txt,
            str(item.get("Código ECL") or ""),
            str(item.get("Responsável") or ""),
            str(item.get("Data/Hora Retirada") or ""),
            str(item.get("Data/Hora Devolução") or ""),
            str(item.get("Status") or ""),
        ])

    table = Table(
        data, repeatRows=1,
        colWidths=[27*mm, 11*mm, 20*mm, 38*mm, 31*mm, 31*mm, 20*mm]
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#17324D")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 7.4),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#BCC5CD")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F7F9")]),
        ("LEFTPADDING", (0,0), (-1,-1), 3),
        ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(table)
    doc.build(story)
    return path



def generate_manual_keys_pdf(keys) -> Path:
    """Gera o mapa manual atualizado das Chaves do ECL, incluindo cópias."""
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = REPORT_DIR / "Controle_Chaves_Polo_Caruaru.pdf"
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        rightMargin=8 * mm, leftMargin=8 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title="Controle de Chaves - Polo Caruaru",
        author="Claviculário ECL",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "ManualTitle", parent=styles["Title"], alignment=TA_CENTER,
        fontName="Helvetica-Bold", fontSize=15, leading=18,
        textColor=colors.HexColor("#17324D")
    )
    subtitle = ParagraphStyle(
        "ManualSubtitle", parent=styles["BodyText"], alignment=TA_CENTER,
        fontName="Helvetica", fontSize=9.5, leading=12,
        textColor=colors.HexColor("#4C5B68")
    )
    story = []
    if LOGO_PDF_PATH.exists():
        logo = RLImage(str(LOGO_PDF_PATH), width=72 * mm, height=21 * mm)
        logo.hAlign = "CENTER"
        story.extend([logo, Spacer(1, 3 * mm)])
    story.extend([
        Paragraph("CONTROLE DE CHAVES - POLO CARUARU", title),
        Paragraph(f"ATUALIZAÇÃO: {datetime.now().strftime('%d/%m/%Y')}", subtitle),
        Spacer(1, 5 * mm),
    ])
    active_keys = [item for item in keys if str(item.get("Ativa", "SIM")).upper() == "SIM"]
    def number_key(item):
        try: return int(item.get("Número no Claviculário") or 9999)
        except (TypeError, ValueError): return 9999
    active_keys.sort(key=number_key)
    data = [["Nº", "Código ECL", "Descrição", "Categoria/Módulo", "Local", "Cópias"]]
    for item in active_keys:
        try: number_txt = f"{int(item.get('Número no Claviculário')):03d}"
        except (TypeError, ValueError): number_txt = ""
        try: copies = int(item.get("Cópias") or 0)
        except (TypeError, ValueError): copies = 0
        copies_txt = str(copies) if copies > 0 else "Sem cópia"
        data.append([
            number_txt, str(item.get("Código") or ""), str(item.get("Descrição") or ""),
            str(item.get("Categoria/Módulo") or ""), str(item.get("Local") or ""), copies_txt
        ])
    if len(data) == 1:
        data.append(["-", "-", "Nenhuma chave cadastrada", "-", "-", "-"])
    table = Table(data, repeatRows=1, colWidths=[11*mm, 22*mm, 52*mm, 39*mm, 48*mm, 22*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#17324D")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 7.6),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#AEB8C1")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F5F7F9")]),
        ("LEFTPADDING", (0,0), (-1,-1), 3), ("RIGHTPADDING", (0,0), (-1,-1), 3),
        ("TOPPADDING", (0,0), (-1,-1), 4), ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(table)
    story.extend([Spacer(1, 4*mm), Paragraph(f"Total de posições ocupadas: {len(active_keys)}", subtitle)])
    doc.build(story)
    return path

def generate_general_keys_pdf(repo) -> Path:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = REPORT_DIR / f"Relatorio_Geral_Chaves_{stamp}.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=7*mm, leftMargin=7*mm,
                            topMargin=10*mm, bottomMargin=10*mm,
                            title="Relatório Geral de Chaves - Polo Caruaru")
    styles = getSampleStyleSheet()
    title = ParagraphStyle("KeysTitle", parent=styles["Title"], alignment=TA_CENTER,
                           fontName="Helvetica-Bold", fontSize=14, textColor=colors.HexColor("#17324D"))
    small = ParagraphStyle("KeysSmall", parent=styles["BodyText"], alignment=TA_CENTER, fontSize=8.5)
    story=[]
    if LOGO_PDF_PATH.exists():
        logo=RLImage(str(LOGO_PDF_PATH), width=68*mm, height=20*mm); logo.hAlign="CENTER"
        story.extend([logo, Spacer(1,2*mm)])
    story.extend([Paragraph("RELATÓRIO GERAL DE CHAVES - POLO CARUARU", title),
                  Paragraph(f"Atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}", small), Spacer(1,4*mm)])
    rows=[["Nº","Código","Descrição","Módulo","Cópias","Total","Ret.","Disp.","Situação"]]
    keys=repo.list_keys()
    keys.sort(key=lambda k: int(k.get("Número no Claviculário") or 9999))
    for k in keys:
        info=repo.key_stock_info(k.get("Código")) or {"copias":0,"total":1,"retiradas":0,"disponiveis":1,"status":"Disponível"}
        try: n=f"{int(k.get('Número no Claviculário')):03d}"
        except: n=""
        rows.append([n,str(k.get("Código") or ""),str(k.get("Descrição") or ""),str(k.get("Categoria/Módulo") or ""),
                     str(info["copias"]),str(info["total"]),str(info["retiradas"]),str(info["disponiveis"]),str(info["status"])])
    if len(rows)==1: rows.append(["-","-","Nenhuma chave cadastrada","-","-","-","-","-","-"])
    table=Table(rows, repeatRows=1, colWidths=[10*mm,19*mm,42*mm,31*mm,13*mm,12*mm,11*mm,12*mm,35*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#17324D")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),6.8),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#AEB8C1")),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F7F9")]),
    ]))
    story.append(table); doc.build(story); return path

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1360x820")
        self.minsize(1160, 700)
        self.repo = ExcelRepository(DB_FILE)

        self.style = ttk.Style(self)
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass

        self._configure_styles()
        self._build_ui()
        self.show_tab("dashboard")
        self.refresh_all()

    def _configure_styles(self):
        self.style.configure("Treeview", rowheight=29, font=("Segoe UI", 10))
        self.style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
        self.style.configure("TButton", font=("Segoe UI", 10), padding=(10, 7))
        self.style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 8))
        self.style.configure("Danger.TButton", font=("Segoe UI", 10, "bold"), padding=(10, 7))

    def _build_ui(self):
        sidebar = tk.Frame(self, bg="#17324D", width=245)
        self.sidebar = sidebar
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        self.brand_logo = None
        if LOGO_HEADER_PATH.exists():
            try:
                logo_img = PILImage.open(LOGO_HEADER_PATH)
                logo_img.thumbnail((205, 64), PILImage.Resampling.LANCZOS)
                self.brand_logo = ImageTk.PhotoImage(logo_img)
                tk.Label(sidebar, image=self.brand_logo, bg="#17324D", bd=0).pack(pady=(18, 4))
            except Exception as exc:
                log_error(exc)

        tk.Label(
            sidebar, text="CLAVICULÁRIO\nECL", bg="#17324D", fg="white",
            font=("Segoe UI", 18, "bold"), pady=12
        ).pack(fill="x")

        menu_items = [
            ("Dashboard", "dashboard"),
            ("Chaves do ECL", "chaves"),
            ("Mapa 1-200", "mapa"),
            ("Retirada de Chave", "retirada"),
            ("Devolução", "devolucao"),
            ("Protocolos", "protocolos"),
            ("Histórico", "historico"),
            ("Relatórios", "relatorios"),
            ("Google Drive / Backup", "backup_drive"),
        ]
        for text, tab in menu_items:
            tk.Button(
                sidebar, text=text, command=lambda t=tab: self.show_tab(t),
                anchor="w", bg="#17324D", fg="white",
                activebackground="#244A6D", activeforeground="white",
                relief="flat", cursor="hand2",
                font=("Segoe UI", 11), padx=24, pady=11
            ).pack(fill="x")

        tk.Label(
            sidebar, text=f"Versão {APP_VERSION}", bg="#17324D", fg="#BFD0DE",
            font=("Segoe UI", 9)
        ).pack(side="bottom", pady=16)

        body = tk.Frame(self, bg="#F3F6F9")
        body.pack(side="left", fill="both", expand=True)

        top = tk.Frame(body, bg="white", height=68)
        top.pack(fill="x")
        top.pack_propagate(False)
        tk.Label(
            top, text=APP_NAME, bg="white", fg="#17324D",
            font=("Segoe UI", 16, "bold")
        ).pack(side="left", padx=26, pady=19)

        ttk.Button(top, text="Abrir planilha", command=self.open_database).pack(
            side="right", padx=(6, 18), pady=14
        )
        ttk.Button(top, text="Backup agora", command=self.manual_backup).pack(
            side="right", padx=6, pady=14
        )

        self.content = tk.Frame(body, bg="#F3F6F9")
        self.content.pack(fill="both", expand=True, padx=20, pady=18)
        self.frames = {}

        self._build_dashboard()
        self._build_keys()
        self._build_map()
        self._build_withdraw()
        self._build_return()
        self._build_protocols()
        self._build_history()
        self._build_reports()
        self._build_backup_drive()

    def _build_backup_drive(self):
        frame = self._new_frame("backup_drive")
        self.title_label(
            frame,
            "Google Drive / Backup",
            "Backup local, cópia no Google Drive e restauração da base em outro computador."
        )

        card = tk.Frame(frame, bg="white", highlightthickness=1, highlightbackground="#D8E0E7")
        card.pack(fill="x", pady=(6, 12), ipady=12)

        self.drive_status_var = tk.StringVar(value="Google Drive: não configurado")
        tk.Label(
            card, textvariable=self.drive_status_var,
            bg="white", fg="#17324D", font=("Segoe UI", 11, "bold")
        ).pack(anchor="w", padx=20, pady=(14, 5))

        self.drive_path_var = tk.StringVar(value="")
        tk.Label(
            card, textvariable=self.drive_path_var,
            bg="white", fg="#5A6975", font=("Segoe UI", 9),
            wraplength=850, justify="left"
        ).pack(anchor="w", padx=20, pady=(0, 12))

        actions = tk.Frame(card, bg="white")
        actions.pack(fill="x", padx=20, pady=(0, 12))

        ttk.Button(
            actions, text="Conectar / escolher pasta do Google Drive",
            command=self.select_google_drive_folder
        ).pack(side="left", padx=(0, 8))

        ttk.Button(
            actions, text="Backup agora",
            style="Primary.TButton",
            command=self.backup_to_google_drive
        ).pack(side="left", padx=8)

        ttk.Button(
            actions, text="Restaurar backup",
            command=self.restore_from_google_drive
        ).pack(side="left", padx=8)

        ttk.Button(
            actions, text="Abrir backups",
            command=self.open_google_backup_folder
        ).pack(side="left", padx=8)

        help_box = tk.Frame(frame, bg="#EEF4F8", highlightthickness=1, highlightbackground="#D6E1E8")
        help_box.pack(fill="x")
        tk.Label(
            help_box,
            text=(
                "Como funciona:\n"
                "1. Instale e entre no Google Drive para computador.\n"
                "2. Clique em 'Conectar / escolher pasta' e selecione Meu Drive (ou uma pasta dentro dele).\n"
                "3. O Claviculário cria automaticamente 'Claviculario ECL\\Backups'.\n"
                "4. Em outro computador, conecte o mesmo Google Drive e use 'Restaurar backup'.\n\n"
                "Antes de uma restauração, o sistema cria automaticamente um backup local de segurança."
            ),
            bg="#EEF4F8", fg="#34495A", justify="left",
            font=("Segoe UI", 10), padx=18, pady=16
        ).pack(anchor="w")

        self.refresh_drive_status()

    def refresh_drive_status(self):
        cfg = load_app_config()
        raw = str(cfg.get("google_drive_folder") or "").strip()
        if raw:
            path = Path(raw)
            if path.exists():
                self.drive_status_var.set("Google Drive: conectado")
                self.drive_path_var.set(f"Pasta selecionada: {path}")
                return
            self.drive_status_var.set("Google Drive: pasta configurada não encontrada")
            self.drive_path_var.set(raw)
        else:
            self.drive_status_var.set("Google Drive: não configurado")
            self.drive_path_var.set("Selecione a pasta sincronizada do Google Drive neste computador.")

    def select_google_drive_folder(self):
        initial = None
        cfg = load_app_config()
        configured = str(cfg.get("google_drive_folder") or "").strip()
        if configured and Path(configured).exists():
            initial = configured

        selected = filedialog.askdirectory(
            title="Selecione Meu Drive ou uma pasta sincronizada do Google Drive",
            initialdir=initial or str(Path.home())
        )
        if not selected:
            return
        cfg["google_drive_folder"] = selected
        save_app_config(cfg)
        target = google_backup_dir()
        if target:
            target.mkdir(parents=True, exist_ok=True)
        self.refresh_drive_status()
        messagebox.showinfo(
            "Google Drive",
            "Pasta conectada com sucesso.\n\n"
            "Os próximos backups poderão ser enviados para o Google Drive."
        )

    def backup_to_google_drive(self):
        try:
            target = google_backup_dir()
            if not target:
                messagebox.showwarning(
                    "Google Drive",
                    "Primeiro conecte/selecione a pasta do Google Drive."
                )
                return
            if not target.parent.parent.exists():
                messagebox.showwarning(
                    "Google Drive",
                    "A pasta configurada não está disponível neste computador."
                )
                return
            path = create_full_backup_zip(target)
            messagebox.showinfo(
                "Backup concluído",
                f"Backup enviado para o Google Drive com sucesso:\n\n{path.name}"
            )
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro no backup", str(exc))

    def restore_from_google_drive(self):
        try:
            target = google_backup_dir()
            initial = str(target) if target and target.exists() else str(Path.home())
            selected = filedialog.askopenfilename(
                title="Selecione o backup do Claviculário ECL",
                initialdir=initial,
                filetypes=[("Backup do Claviculário", "*.zip"), ("Todos os arquivos", "*.*")]
            )
            if not selected:
                return

            if not messagebox.askyesno(
                "Restaurar backup",
                "Deseja restaurar este backup?\n\n"
                "A base atual será substituída. Antes disso, o sistema fará um backup local de segurança."
            ):
                return

            restore_backup_zip(Path(selected))
            self.repo.ensure_database()
            self.refresh_all()
            messagebox.showinfo(
                "Restauração concluída",
                "Backup restaurado com sucesso."
            )
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro na restauração", str(exc))

    def open_google_backup_folder(self):
        try:
            target = google_backup_dir()
            if not target:
                messagebox.showwarning("Google Drive", "Google Drive ainda não configurado.")
                return
            target.mkdir(parents=True, exist_ok=True)
            open_file(target)
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    def _new_frame(self, name):
        frame = tk.Frame(self.content, bg="#F3F6F9")
        frame.place(relx=0, rely=0, relwidth=1, relheight=1)
        self.frames[name] = frame
        return frame

    def show_tab(self, name):
        self.frames[name].tkraise()
        try:
            self.refresh_all()
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    def title_label(self, frame, text, subtitle=None):
        tk.Label(frame, text=text, bg="#F3F6F9", fg="#17324D",
                 font=("Segoe UI", 22, "bold")).pack(anchor="w")
        if subtitle:
            tk.Label(frame, text=subtitle, bg="#F3F6F9", fg="#617180",
                     font=("Segoe UI", 10)).pack(anchor="w", pady=(3, 12))

    def _build_dashboard(self):
        frame = self._new_frame("dashboard")
        self.title_label(frame, "Dashboard", "Visão geral do Claviculário ECL")

        cards = tk.Frame(frame, bg="#F3F6F9")
        cards.pack(fill="x")
        self.card_vars = []
        for title in ["Total de Chaves", "Com unidade disponível", "Com retiradas", "Espaços livres"]:
            card = tk.Frame(cards, bg="white", highlightthickness=1, highlightbackground="#D8E0E7")
            card.pack(side="left", fill="x", expand=True, padx=(0, 12), ipady=10)
            tk.Label(card, text=title, bg="white", fg="#607080",
                     font=("Segoe UI", 10)).pack(anchor="w", padx=16, pady=(10, 0))
            var = tk.StringVar(value="0")
            self.card_vars.append(var)
            tk.Label(card, textvariable=var, bg="white", fg="#17324D",
                     font=("Segoe UI", 27, "bold")).pack(anchor="w", padx=16, pady=(2, 8))

        lower = tk.Frame(frame, bg="#F3F6F9")
        lower.pack(fill="both", expand=True, pady=(18, 0))

        chart_box = tk.Frame(lower, bg="white", highlightthickness=1, highlightbackground="#D8E0E7")
        chart_box.pack(side="left", fill="both", expand=True, padx=(0, 12))
        tk.Label(chart_box, text="Chaves por categoria/módulo", bg="white", fg="#17324D",
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=16, pady=14)
        self.chart_canvas = tk.Canvas(chart_box, bg="white", highlightthickness=0)
        self.chart_canvas.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.chart_canvas.bind("<Configure>", lambda e: self.draw_chart())

        side = tk.Frame(lower, bg="white", width=340, highlightthickness=1, highlightbackground="#D8E0E7")
        side.pack(side="right", fill="y")
        side.pack_propagate(False)
        tk.Label(side, text="Operação", bg="white", fg="#17324D",
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=18, pady=(18, 8))
        tk.Label(
            side,
            text=(
                "Fluxo recomendado:\n\n"
                "1. Cadastre a Chave do ECL\n"
                "2. Registre a retirada\n"
                "3. Imprima o protocolo\n"
                "4. Colha as assinaturas\n"
                "5. Registre a devolução\n\n"
                "Todo o histórico fica preservado.\n\n"
                "Use o Mapa 1–200 para localizar\nrapidamente cada posição física."
            ),
            justify="left", bg="white", fg="#4C5B68", font=("Segoe UI", 10)
        ).pack(anchor="w", padx=18)

    def _build_keys(self):
        frame = self._new_frame("chaves")
        top = tk.Frame(frame, bg="#F3F6F9"); top.pack(fill="x")
        tk.Label(top, text="Chaves do ECL", bg="#F3F6F9", fg="#17324D", font=("Segoe UI",22,"bold")).pack(side="left")
        ttk.Button(top, text="+ Nova Chave", style="Primary.TButton", command=self.open_key_form).pack(side="right")
        search=tk.Frame(frame,bg="#F3F6F9"); search.pack(fill="x",pady=12)
        tk.Label(search,text="Pesquisar:",bg="#F3F6F9").pack(side="left")
        self.key_search=tk.StringVar(); ent=ttk.Entry(search,textvariable=self.key_search,width=44); ent.pack(side="left",padx=8)
        ent.bind("<KeyRelease>",lambda e:self.refresh_keys())
        cols=("numero","codigo","descricao","categoria","local","copias","disponiveis","status")
        self.key_tree=ttk.Treeview(frame,columns=cols,show="headings")
        config=[("numero","Nº",58),("codigo","Código",95),("descricao","Descrição",230),("categoria","Categoria/Módulo",165),
                ("local","Local",165),("copias","Cópias",70),("disponiveis","Disponíveis",80),("status","Situação",150)]
        for c,h,w in config:
            self.key_tree.heading(c,text=h); self.key_tree.column(c,width=w,anchor="w")
        self.key_tree.pack(fill="both",expand=True)
        bar=tk.Frame(frame,bg="#F3F6F9"); bar.pack(fill="x",pady=10)
        ttk.Button(bar,text="Editar",command=self.edit_selected_key).pack(side="left")
        ttk.Button(bar,text="Marcar como Livre",command=self.free_selected_key).pack(side="left",padx=8)

    def _build_map(self):
        frame = self._new_frame("mapa")
        self.title_label(
            frame, "Mapa do Claviculário 1–200",
            "Passe o mouse sobre uma posição para ver os dados. Clique somente quando quiser abrir as opções."
        )

        legend = tk.Frame(frame, bg="#F3F6F9")
        legend.pack(fill="x", pady=(0, 8))
        for label, color in [
            ("Disponível", "#D8F3DC"),
            ("Parcial", "#FFF0B3"),
            ("Sem unidade", "#FFD6D6"),
            ("Livre", "#FFFFFF"),
        ]:
            tk.Label(legend, text="  ", bg=color, relief="solid", bd=1).pack(side="left", padx=(0, 4))
            tk.Label(legend, text=label, bg="#F3F6F9", fg="#465563").pack(side="left", padx=(0, 16))

        controls = tk.Frame(frame, bg="#F3F6F9")
        controls.pack(fill="x", pady=(0, 8))
        tk.Label(controls, text="Filtrar módulo:", bg="#F3F6F9").pack(side="left")
        self.map_category = tk.StringVar(value="Todos")
        combo = ttk.Combobox(
            controls, textvariable=self.map_category,
            values=["Todos"] + CATEGORIAS, state="readonly", width=24
        )
        combo.pack(side="left", padx=8)
        combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_map())
        ttk.Button(controls, text="Atualizar", command=self.refresh_map).pack(side="left")

        container = tk.Frame(frame, bg="white", highlightthickness=1, highlightbackground="#D8E0E7")
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        self.map_inner = tk.Frame(canvas, bg="white")
        self.map_inner.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=self.map_inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")

        self.map_buttons = {}

    def show_map_position(self, number):
        found = None
        for item in self.repo.list_keys():
            try:
                if int(item.get("Número no Claviculário")) == int(number):
                    found = item; break
            except (TypeError, ValueError):
                continue
        if not found:
            if messagebox.askyesno("Posição livre", f"Posição {int(number):03d} está LIVRE.\n\nDeseja cadastrar uma chave nesta posição?"):
                self.open_key_form(preset_number=number)
            return
        stock = self.repo.key_stock_info(found.get("Código")) or {}
        win=tk.Toplevel(self); win.title(f"Posição {int(number):03d}"); win.geometry("520x390"); win.resizable(False,False); win.transient(self); win.grab_set()
        tk.Label(win,text=f"POSIÇÃO {int(number):03d}",font=("Segoe UI",16,"bold"),fg="#17324D").pack(pady=(22,12))
        copies=stock.get("copias",0); copy_text=str(copies) if copies else "Sem cópia"
        info=(f"Chave do ECL: {found.get('Código')}\nDescrição: {found.get('Descrição')}\n"
              f"Categoria/Módulo: {found.get('Categoria/Módulo')}\nLocal: {found.get('Local') or '-'}\n"
              f"Cópias: {copy_text}\nTotal físico: {stock.get('total',1)}\n"
              f"Retiradas: {stock.get('retiradas',0)} | Disponíveis: {stock.get('disponiveis',1)}\n"
              f"Situação: {stock.get('status','Disponível')}")
        tk.Label(win,text=info,justify="left",font=("Segoe UI",10),padx=20,pady=10).pack(fill="x")
        actions=tk.Frame(win); actions.pack(pady=18)
        ttk.Button(actions,text="Editar",command=lambda:(win.destroy(),self.open_key_form(found))).pack(side="left",padx=5)
        if stock.get("disponiveis",0)>0:
            ttk.Button(actions,text="Retirar",command=lambda:(win.destroy(),self.start_withdraw_for_code(found.get("Código")))).pack(side="left",padx=5)
        ttk.Button(actions,text="Marcar como Livre",command=lambda:(win.destroy(),self.free_key_item(found))).pack(side="left",padx=5)

    def _build_withdraw(self):
        frame = self._new_frame("retirada")
        self.title_label(frame, "Retirada de Chave", "Registre a saída e gere o protocolo automaticamente")

        box = tk.Frame(frame, bg="white", highlightthickness=1, highlightbackground="#D8E0E7")
        box.pack(fill="both", expand=True, pady=(4, 0))

        form = tk.Frame(box, bg="white")
        form.pack(fill="x", padx=24, pady=20)

        self.withdraw_vars = {
            "codigo": tk.StringVar(),
            "responsavel": tk.StringVar(),
            "setor": tk.StringVar(),
            "motivo": tk.StringVar(),
            "observacao": tk.StringVar(),
        }

        labels = [
            ("Chave do ECL", "codigo"),
            ("Responsável", "responsavel"),
            ("Setor", "setor"),
            ("Motivo da retirada", "motivo"),
            ("Observações", "observacao"),
        ]

        for r, (label, key) in enumerate(labels):
            tk.Label(form, text=label, bg="white", fg="#273848",
                     font=("Segoe UI", 10, "bold")).grid(row=r, column=0, sticky="w", padx=(0, 15), pady=8)
            if key == "codigo":
                self.withdraw_combo = ttk.Combobox(
                    form, textvariable=self.withdraw_vars[key], state="readonly", width=55
                )
                self.withdraw_combo.grid(row=r, column=1, sticky="ew", pady=8)
                self.withdraw_combo.bind("<<ComboboxSelected>>", lambda e: self.update_withdraw_key_info())
            else:
                ttk.Entry(form, textvariable=self.withdraw_vars[key], width=58).grid(
                    row=r, column=1, sticky="ew", pady=8
                )
        form.columnconfigure(1, weight=1)

        self.withdraw_info = tk.StringVar(value="Selecione uma Chave do ECL disponível.")
        tk.Label(
            box, textvariable=self.withdraw_info, justify="left", anchor="w",
            bg="#F4F7FA", fg="#17324D", font=("Segoe UI", 10),
            padx=16, pady=12
        ).pack(fill="x", padx=24, pady=(0, 12))

        actions = tk.Frame(box, bg="white")
        actions.pack(fill="x", padx=24, pady=(4, 22))
        ttk.Button(
            actions, text="CONFIRMAR RETIRADA E GERAR PROTOCOLO",
            style="Primary.TButton", command=self.confirm_withdrawal
        ).pack(side="left")
        ttk.Button(actions, text="Limpar", command=self.clear_withdraw_form).pack(side="left", padx=10)

    def _build_return(self):
        frame = self._new_frame("devolucao")
        self.title_label(frame, "Devolução de Chave", "Somente retiradas em aberto aparecem nesta lista")

        cols = ("protocolo", "numero", "codigo", "descricao", "responsavel", "setor", "retirada")
        self.return_tree = ttk.Treeview(frame, columns=cols, show="headings")
        config = [
            ("protocolo", "Protocolo", 135),
            ("numero", "Nº", 65),
            ("codigo", "Código", 100),
            ("descricao", "Descrição", 280),
            ("responsavel", "Responsável", 220),
            ("setor", "Setor", 170),
            ("retirada", "Retirada", 170),
        ]
        for c, h, w in config:
            self.return_tree.heading(c, text=h)
            self.return_tree.column(c, width=w, anchor="w")
        self.return_tree.pack(fill="both", expand=True, pady=(4, 0))

        bar = tk.Frame(frame, bg="#F3F6F9")
        bar.pack(fill="x", pady=10)
        ttk.Button(
            bar, text="DEVOLVER CHAVE SELECIONADA",
            style="Primary.TButton", command=self.confirm_return
        ).pack(side="left")

    def _build_protocols(self):
        frame = self._new_frame("protocolos")
        self.title_label(frame, "Protocolos", "Consulte e reabra os protocolos gerados")

        s = tk.Frame(frame, bg="#F3F6F9")
        s.pack(fill="x", pady=(0, 10))
        tk.Label(s, text="Pesquisar:", bg="#F3F6F9").pack(side="left")
        self.protocol_search = tk.StringVar()
        ent = ttk.Entry(s, textvariable=self.protocol_search, width=42)
        ent.pack(side="left", padx=8)
        ent.bind("<KeyRelease>", lambda e: self.refresh_protocols())

        cols = ("protocolo", "numero", "codigo", "descricao", "responsavel", "retirada", "status")
        self.protocol_tree = ttk.Treeview(frame, columns=cols, show="headings")
        config = [
            ("protocolo", "Protocolo", 130),
            ("numero", "Nº", 65),
            ("codigo", "Código", 95),
            ("descricao", "Descrição", 260),
            ("responsavel", "Responsável", 220),
            ("retirada", "Retirada", 170),
            ("status", "Status", 115),
        ]
        for c, h, w in config:
            self.protocol_tree.heading(c, text=h)
            self.protocol_tree.column(c, width=w, anchor="w")
        self.protocol_tree.pack(fill="both", expand=True)

        bar = tk.Frame(frame, bg="#F3F6F9")
        bar.pack(fill="x", pady=10)
        ttk.Button(bar, text="Abrir PDF", command=self.open_selected_protocol).pack(side="left")
        ttk.Button(bar, text="Gerar/Regerar PDF", command=self.regenerate_selected_protocol).pack(side="left", padx=8)
        ttk.Button(bar, text="Abrir pasta de protocolos", command=lambda: open_file(PROTOCOL_DIR)).pack(side="left")

    def _build_history(self):
        frame = self._new_frame("historico")
        self.title_label(frame, "Histórico", "Registro permanente das retiradas e devoluções")

        s = tk.Frame(frame, bg="#F3F6F9")
        s.pack(fill="x", pady=(0, 10))
        tk.Label(s, text="Pesquisar:", bg="#F3F6F9").pack(side="left")
        self.history_search = tk.StringVar()
        ent = ttk.Entry(s, textvariable=self.history_search, width=45)
        ent.pack(side="left", padx=8)
        ent.bind("<KeyRelease>", lambda e: self.refresh_history())

        cols = ("protocolo", "numero", "codigo", "responsavel", "retirada", "devolucao", "status")
        self.history_tree = ttk.Treeview(frame, columns=cols, show="headings")
        config = [
            ("protocolo", "Protocolo", 135),
            ("numero", "Nº", 65),
            ("codigo", "Código", 100),
            ("responsavel", "Responsável", 250),
            ("retirada", "Retirada", 180),
            ("devolucao", "Devolução", 180),
            ("status", "Status", 120),
        ]
        for c, h, w in config:
            self.history_tree.heading(c, text=h)
            self.history_tree.column(c, width=w, anchor="w")
        self.history_tree.pack(fill="both", expand=True)

    def _build_reports(self):
        frame = self._new_frame("relatorios")
        self.title_label(
            frame, "Relatórios",
            "Filtre movimentações por período, módulo, status ou responsável e gere um PDF."
        )

        filters = tk.Frame(frame, bg="white", highlightthickness=1, highlightbackground="#D8E0E7")
        filters.pack(fill="x", pady=(0, 10), ipady=8)

        self.report_start = tk.StringVar()
        self.report_end = tk.StringVar()
        self.report_category = tk.StringVar(value="Todos")
        self.report_status = tk.StringVar(value="Todos")
        self.report_search = tk.StringVar()

        row = tk.Frame(filters, bg="white")
        row.pack(fill="x", padx=14, pady=8)

        tk.Label(row, text="Data inicial:", bg="white").pack(side="left")
        ttk.Entry(row, textvariable=self.report_start, width=13).pack(side="left", padx=(5, 12))
        tk.Label(row, text="Data final:", bg="white").pack(side="left")
        ttk.Entry(row, textvariable=self.report_end, width=13).pack(side="left", padx=(5, 12))

        tk.Label(row, text="Módulo:", bg="white").pack(side="left")
        ttk.Combobox(
            row, textvariable=self.report_category,
            values=["Todos"] + CATEGORIAS, state="readonly", width=20
        ).pack(side="left", padx=(5, 12))

        tk.Label(row, text="Status:", bg="white").pack(side="left")
        ttk.Combobox(
            row, textvariable=self.report_status,
            values=["Todos", "Retirada", "Devolvida"], state="readonly", width=13
        ).pack(side="left", padx=(5, 12))

        row2 = tk.Frame(filters, bg="white")
        row2.pack(fill="x", padx=14, pady=(0, 8))
        tk.Label(row2, text="Pesquisar:", bg="white").pack(side="left")
        ent = ttk.Entry(row2, textvariable=self.report_search, width=42)
        ent.pack(side="left", padx=6)
        ent.bind("<Return>", lambda e: self.apply_report_filters())

        ttk.Button(row2, text="Aplicar filtros", command=self.apply_report_filters).pack(side="left", padx=6)
        ttk.Button(row2, text="Limpar", command=self.clear_report_filters).pack(side="left", padx=6)
        ttk.Button(
            row2, text="Gerar PDF do relatório", style="Primary.TButton",
            command=self.export_report_pdf
        ).pack(side="right")
        ttk.Button(
            row2, text="Relatório Manual de Chaves", style="Primary.TButton",
            command=self.export_manual_keys_pdf
        ).pack(side="right", padx=(0, 8))
        ttk.Button(
            row2, text="Relatório Geral de Chaves",
            command=self.export_general_keys_pdf
        ).pack(side="right", padx=(0, 8))

        self.report_total = tk.StringVar(value="0 registros")
        tk.Label(
            frame, textvariable=self.report_total, bg="#F3F6F9", fg="#17324D",
            font=("Segoe UI", 10, "bold")
        ).pack(anchor="w", pady=(0, 6))

        cols = ("protocolo", "numero", "codigo", "categoria", "responsavel", "retirada", "devolucao", "status")
        self.report_tree = ttk.Treeview(frame, columns=cols, show="headings")
        config = [
            ("protocolo", "Protocolo", 125),
            ("numero", "Nº", 60),
            ("codigo", "Chave", 90),
            ("categoria", "Módulo", 150),
            ("responsavel", "Responsável", 190),
            ("retirada", "Retirada", 155),
            ("devolucao", "Devolução", 155),
            ("status", "Status", 100),
        ]
        for c, h, w in config:
            self.report_tree.heading(c, text=h)
            self.report_tree.column(c, width=w, anchor="w")
        self.report_tree.pack(fill="both", expand=True)

        self.current_report_rows = []

    def report_filters_dict(self):
        return {
            "inicio": self.report_start.get().strip(),
            "fim": self.report_end.get().strip(),
            "categoria": self.report_category.get().strip(),
            "status": self.report_status.get().strip(),
            "search": self.report_search.get().strip(),
        }

    def apply_report_filters(self):
        try:
            rows = self.repo.filtered_movements(
                self.report_start.get(),
                self.report_end.get(),
                self.report_category.get(),
                self.report_status.get(),
                self.report_search.get(),
            )
            self.current_report_rows = rows
            for child in self.report_tree.get_children():
                self.report_tree.delete(child)

            for item in rows:
                number = item.get("Número no Claviculário")
                number_txt = f"{int(number):03d}" if number not in (None, "") else ""
                self.report_tree.insert("", "end", values=(
                    item.get("Protocolo", ""),
                    number_txt,
                    item.get("Código ECL", ""),
                    item.get("Categoria/Módulo", ""),
                    item.get("Responsável", ""),
                    item.get("Data/Hora Retirada", ""),
                    item.get("Data/Hora Devolução", ""),
                    item.get("Status", ""),
                ))
            self.report_total.set(f"{len(rows)} registro(s)")
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro nos filtros", str(exc))

    def clear_report_filters(self):
        self.report_start.set("")
        self.report_end.set("")
        self.report_category.set("Todos")
        self.report_status.set("Todos")
        self.report_search.set("")
        self.apply_report_filters()

    def refresh_manual_keys_pdf(self, silent=True):
        """Atualiza o mapa manual; em modo silencioso não abre janelas."""
        try:
            return generate_manual_keys_pdf(self.repo.list_keys(include_inactive=True))
        except Exception as exc:
            log_error(exc)
            if not silent:
                messagebox.showerror("Relatório Manual", str(exc))
            return None

    def export_manual_keys_pdf(self):
        path = self.refresh_manual_keys_pdf(silent=False)
        if not path:
            return
        if messagebox.askyesno(
            "Relatório Manual de Chaves",
            f"PDF atualizado com sucesso:\n{path.name}\n\nDeseja abrir agora?"
        ):
            open_file(path)

    def export_general_keys_pdf(self):
        try:
            path=generate_general_keys_pdf(self.repo)
            if messagebox.askyesno("Relatório Geral de Chaves",f"Relatório gerado com sucesso:\n{path.name}\n\nDeseja abrir agora?"):
                open_file(path)
        except Exception as exc:
            log_error(exc); messagebox.showerror("Erro",str(exc))

    def export_report_pdf(self):
        try:
            self.apply_report_filters()
            if not self.current_report_rows:
                messagebox.showwarning("Relatório", "Nenhum registro encontrado para os filtros selecionados.")
                return
            path = generate_movement_report_pdf(self.current_report_rows, self.report_filters_dict())
            if messagebox.askyesno(
                "Relatório gerado",
                f"Relatório gerado com sucesso:\n{path.name}\n\nDeseja abrir agora?"
            ):
                open_file(path)
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    def _build_inactive(self):
        frame = self._new_frame("inativas")
        self.title_label(frame, "Chaves Inativas", "Mantidas para preservar rastreabilidade e histórico")

        cols = ("numero", "codigo", "descricao", "categoria", "local", "motivo")
        self.inactive_tree = ttk.Treeview(frame, columns=cols, show="headings")
        config = [
            ("numero", "Nº", 70),
            ("codigo", "Código", 105),
            ("descricao", "Descrição", 300),
            ("categoria", "Categoria/Módulo", 210),
            ("local", "Local", 230),
            ("motivo", "Motivo", 240),
        ]
        for c, h, w in config:
            self.inactive_tree.heading(c, text=h)
            self.inactive_tree.column(c, width=w, anchor="w")
        self.inactive_tree.pack(fill="both", expand=True)

        bar = tk.Frame(frame, bg="#F3F6F9")
        bar.pack(fill="x", pady=10)
        ttk.Button(bar, text="Restaurar Chave", command=self.restore_selected).pack(side="left")

    # ---------- AÇÕES GERAIS ----------

    def open_database(self):
        try:
            open_file(DB_FILE)
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    def manual_backup(self):
        try:
            path = create_full_backup_zip()
            messagebox.showinfo("Backup", f"Backup local criado com sucesso:\n\n{path.name}")
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    # ---------- CHAVES ----------

    def open_key_form(self, item=None, preset_number=None):
        win=tk.Toplevel(self); win.title("Cadastro de Chave do ECL"); win.geometry("555x625"); win.resizable(False,False); win.transient(self); win.grab_set()
        mapping={"Número no Claviculário":"Número no Claviculário","Código ECL":"Código","Descrição":"Descrição","Local":"Local","Cópias":"Cópias","Observação":"Observação"}
        fields={}
        labels=["Número no Claviculário","Código ECL","Descrição","Categoria/Módulo","Local","Cópias","Observação"]
        for i,label in enumerate(labels):
            tk.Label(win,text=label,font=("Segoe UI",10,"bold")).pack(anchor="w",padx=26,pady=(16 if i==0 else 7,3))
            if label=="Categoria/Módulo":
                var=tk.StringVar(value=(item or {}).get("Categoria/Módulo",CATEGORIAS[0])); widget=ttk.Combobox(win,textvariable=var,values=CATEGORIAS,state="readonly")
            else:
                default=(item or {}).get(mapping[label],"")
                if label=="Número no Claviculário" and not item and preset_number is not None: default=str(preset_number)
                if label=="Cópias" and default in (None,""): default="0"
                var=tk.StringVar(value=default); widget=ttk.Entry(win,textvariable=var)
            fields[label]=var; widget.pack(fill="x",padx=26)
        tk.Label(win,text="Cópias = unidades extras além da chave principal. Se não houver, informe 0.",fg="#607080",font=("Segoe UI",8)).pack(anchor="w",padx=26,pady=(3,0))
        def save():
            try:
                args=[fields["Número no Claviculário"].get(),fields["Código ECL"].get(),fields["Descrição"].get(),fields["Categoria/Módulo"].get(),fields["Local"].get(),fields["Cópias"].get(),fields["Observação"].get()]
                if item: self.repo.update_key(item["Código"],*args)
                else: self.repo.add_key(*args)
                win.destroy(); self.refresh_all(); self.refresh_manual_keys_pdf(silent=True)
            except Exception as exc:
                log_error(exc); messagebox.showerror("Não foi possível salvar",str(exc),parent=win)
        ttk.Button(win,text="SALVAR",style="Primary.TButton",command=save).pack(pady=20)

    def selected_active_key(self):
        selected = self.key_tree.selection()
        if not selected:
            raise ValueError("Selecione uma Chave do ECL.")
        code = self.key_tree.item(selected[0], "values")[1]
        for item in self.repo.list_keys():
            if str(item.get("Código")) == str(code):
                return item
        raise ValueError("Registro não encontrado.")

    def edit_selected_key(self):
        try:
            self.open_key_form(self.selected_active_key())
        except Exception as exc:
            messagebox.showwarning("Atenção", str(exc))

    def inactivate_selected(self):
        # Compatibilidade: a ação antiga agora libera a posição.
        self.free_selected_key()

    def free_selected_key(self):
        try:
            item=self.selected_active_key()
        except Exception as exc:
            messagebox.showwarning("Atenção",str(exc)); return
        self.free_key_item(item)

    def free_key_item(self, item):
        stock=self.repo.key_stock_info(item.get("Código")) or {"retiradas":0}
        extra=""
        if stock.get("retiradas",0)>0:
            extra=(f"\n\nATENÇÃO: existem {stock['retiradas']} unidade(s) atualmente retiradas. "
                   "Elas serão registradas no histórico como entrega definitiva / posição liberada.")
        if not messagebox.askyesno("Marcar como Livre",
            f"Deseja liberar a posição {int(item.get('Número no Claviculário')):03d}?\n\n"
            f"{item.get('Código')} - {item.get('Descrição')}\n\n"
            "O número ficará disponível para outra chave e o histórico será preservado."+extra):
            return
        try:
            self.repo.free_key(item.get("Código")); self.refresh_all(); self.refresh_manual_keys_pdf(silent=True)
            messagebox.showinfo("Espaço livre","Posição liberada com sucesso.")
        except Exception as exc:
            log_error(exc); messagebox.showerror("Erro",str(exc))

    def start_withdraw_for_code(self, codigo):
        self.show_tab("retirada")
        for value in self.withdraw_combo["values"]:
            if str(value).endswith(f"| {codigo}"):
                self.withdraw_vars["codigo"].set(value); self.update_withdraw_key_info(); break

    def restore_selected(self):
        selected = self.inactive_tree.selection()
        if not selected:
            messagebox.showwarning("Atenção", "Selecione uma Chave do ECL inativa.")
            return
        code = self.inactive_tree.item(selected[0], "values")[1]
        if not messagebox.askyesno("Restaurar", f"Deseja restaurar a chave {code}?"):
            return
        try:
            self.repo.restore_key(code)
            self.refresh_all()
            self.refresh_manual_keys_pdf(silent=True)
            messagebox.showinfo("Concluído", "Chave restaurada para o cadastro ativo.")
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    # ---------- RETIRADA ----------

    def update_withdraw_key_info(self):
        selection=self.withdraw_vars["codigo"].get(); code=selection.split("|",1)[1].strip() if "|" in selection else selection.strip()
        for item in self.repo.list_keys():
            if str(item.get("Código"))==code:
                numero=item.get("Número no Claviculário"); numero_txt=f"{int(numero):03d}" if numero not in (None,"") else "-"
                stock=self.repo.key_stock_info(code) or {}
                copies=stock.get("copias",0); copy_txt=str(copies) if copies else "Sem cópia"
                alert="" if stock.get("disponiveis",0)>0 else "  |  ALERTA: SEM CHAVE/CÓPIA DISPONÍVEL"
                self.withdraw_info.set(
                    f"Nº {numero_txt} | {item.get('Código')} - {item.get('Descrição')}\n"
                    f"Categoria/Módulo: {item.get('Categoria/Módulo')} | Local: {item.get('Local') or '-'}\n"
                    f"Cópias: {copy_txt} | Total físico: {stock.get('total',1)} | Retiradas: {stock.get('retiradas',0)} | Disponíveis: {stock.get('disponiveis',1)}{alert}"
                ); return
        self.withdraw_info.set("Selecione uma Chave do ECL com unidade disponível.")

    def clear_withdraw_form(self):
        for var in self.withdraw_vars.values():
            var.set("")
        self.withdraw_info.set("Selecione uma Chave do ECL disponível.")

    def confirm_withdrawal(self):
        selection = self.withdraw_vars["codigo"].get().strip()
        code = selection.split("|", 1)[1].strip() if "|" in selection else selection
        if not code:
            messagebox.showwarning("Atenção", "Selecione a Chave do ECL.")
            return

        preview = (
            f"Chave: {code}\n"
            f"Responsável: {self.withdraw_vars['responsavel'].get().strip()}\n"
            f"Setor: {self.withdraw_vars['setor'].get().strip()}\n\n"
            "Confirmar a retirada e gerar o protocolo?"
        )
        if not messagebox.askyesno("Confirmar retirada", preview):
            return

        try:
            data = self.repo.withdraw_key(
                code,
                self.withdraw_vars["responsavel"].get(),
                self.withdraw_vars["setor"].get(),
                self.withdraw_vars["motivo"].get(),
                self.withdraw_vars["observacao"].get(),
            )
            pdf_path = generate_protocol_pdf(data)
            self.repo.set_protocol_pdf(data["protocolo"], pdf_path)

            self.refresh_all()
            self.clear_withdraw_form()

            if messagebox.askyesno(
                "Retirada registrada",
                f"Retirada registrada com sucesso.\n\n"
                f"Protocolo: {data['protocolo']}\n\n"
                "Deseja abrir o protocolo em PDF agora?"
            ):
                open_file(pdf_path)
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Não foi possível registrar a retirada", str(exc))

    # ---------- DEVOLUÇÃO ----------

    def confirm_return(self):
        selected = self.return_tree.selection()
        if not selected:
            messagebox.showwarning("Atenção", "Selecione uma retirada em aberto.")
            return

        values = self.return_tree.item(selected[0], "values")
        protocol, number, code, desc, responsible = values[0], values[1], values[2], values[3], values[4]
        if not messagebox.askyesno(
            "Confirmar devolução",
            f"Protocolo: {protocol}\nNº no Claviculário: {number}\nChave: {code} - {desc}\n"
            f"Responsável: {responsible}\n\n"
            "Confirmar a devolução desta Chave do ECL?"
        ):
            return

        try:
            movement = self.repo.return_key(protocol)
            self.refresh_all()
            messagebox.showinfo(
                "Devolução concluída",
                f"Chave {movement.get('Código ECL')} devolvida com sucesso.\n"
                f"Data/Hora: {movement.get('Data/Hora Devolução')}"
            )
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    # ---------- PROTOCOLOS ----------

    def selected_protocol_number(self):
        selected = self.protocol_tree.selection()
        if not selected:
            raise ValueError("Selecione um protocolo.")
        return str(self.protocol_tree.item(selected[0], "values")[0])

    def open_selected_protocol(self):
        try:
            protocol = self.selected_protocol_number()
            data = self.repo.get_protocol(protocol)
            if not data:
                raise ValueError("Protocolo não encontrado.")
            pdf = str(data.get("Arquivo PDF") or "").strip()
            if not pdf or not Path(pdf).exists():
                raise ValueError("O PDF deste protocolo ainda não existe. Use 'Gerar/Regerar PDF'.")
            open_file(Path(pdf))
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    def regenerate_selected_protocol(self):
        try:
            protocol = self.selected_protocol_number()
            p = self.repo.get_protocol(protocol)
            if not p:
                raise ValueError("Protocolo não encontrado.")

            data = {
                "protocolo": p.get("Protocolo"),
                "numero": p.get("Número no Claviculário"),
                "codigo": p.get("Código ECL"),
                "descricao": p.get("Descrição"),
                "categoria": p.get("Categoria/Módulo"),
                "local": "",
                "responsavel": p.get("Responsável"),
                "setor": p.get("Setor"),
                "motivo": p.get("Motivo"),
                "observacao": p.get("Observação"),
                "retirada": p.get("Data/Hora Retirada"),
            }
            for k in self.repo.list_keys(include_inactive=True):
                if str(k.get("Código")) == str(data["codigo"]):
                    data["local"] = k.get("Local") or ""
                    break

            pdf = generate_protocol_pdf(data)
            self.repo.set_protocol_pdf(protocol, pdf)
            self.refresh_protocols()
            open_file(pdf)
        except Exception as exc:
            log_error(exc)
            messagebox.showerror("Erro", str(exc))

    # ---------- ATUALIZAÇÕES DAS TELAS ----------


    def map_hover_text(self, number, item):
        if not item:
            return f"POSIÇÃO {int(number):03d}\nLIVRE\nClique para cadastrar uma chave."

        stock = self.repo.key_stock_info(item.get("Código")) or {}
        copies = stock.get("copias", 0)
        copy_text = str(copies) if copies else "Sem cópia"
        return (
            f"POSIÇÃO {int(number):03d}\n"
            f"{item.get('Código')} - {item.get('Descrição')}\n"
            f"Módulo: {item.get('Categoria/Módulo')}\n"
            f"Local: {item.get('Local') or '-'}\n"
            f"Cópias: {copy_text}\n"
            f"Disponíveis: {stock.get('disponiveis', 1)} | Retiradas: {stock.get('retiradas', 0)}\n"
            f"Situação: {stock.get('status', 'Disponível')}\n\n"
            "Clique para abrir as opções."
        )

    def refresh_map(self):
        if not hasattr(self,"map_inner"): return
        for child in self.map_inner.winfo_children(): child.destroy()
        selected_category=self.map_category.get() if hasattr(self,"map_category") else "Todos"
        positions={}
        for item in self.repo.list_keys():
            try: positions[int(item.get("Número no Claviculário"))]=item
            except (TypeError,ValueError): continue
        for n in range(1,201):
            item=positions.get(n); bg="#FFFFFF"; fg="#17324D"
            if item:
                cat=item.get("Categoria/Módulo")
                if selected_category!="Todos" and cat!=selected_category:
                    bg="#F5F6F7"; fg="#A5ADB4"
                else:
                    stock=self.repo.key_stock_info(item.get("Código")) or {}
                    if stock.get("disponiveis",0)<=0: bg="#FFD6D6"
                    elif stock.get("retiradas",0)>0: bg="#FFF0B3"
                    else: bg="#D8F3DC"
            btn=tk.Button(self.map_inner,text=f"{n:03d}",width=6,height=2,bg=bg,fg=fg,relief="solid",bd=1,cursor="hand2",font=("Segoe UI",9,"bold"),command=lambda num=n:self.show_map_position(num))
            btn.grid(row=(n-1)//10,column=(n-1)%10,padx=3,pady=3,sticky="nsew")
            MapTooltip(btn, lambda num=n, current=item: self.map_hover_text(num, current))
            self.map_buttons[n]=btn
        for col in range(10): self.map_inner.grid_columnconfigure(col,weight=1)

    def refresh_all(self):
        self.refresh_dashboard()
        self.refresh_keys()
        self.refresh_map()
        self.refresh_withdraw_combo()
        self.refresh_returns()
        self.refresh_protocols()
        self.refresh_history()
        if hasattr(self, "report_tree"):
            self.apply_report_filters()
        if hasattr(self, "drive_status_var"):
            self.refresh_drive_status()

    def refresh_dashboard(self):
        counts = self.repo.dashboard_counts()
        for var, value in zip(self.card_vars, counts):
            var.set(str(value))
        self.draw_chart()

    def draw_chart(self):
        if not hasattr(self, "chart_canvas"):
            return
        c = self.chart_canvas
        c.delete("all")
        counts = self.repo.category_counts()
        items = list(counts.items())
        c.update_idletasks()
        w = max(c.winfo_width(), 560)
        h = max(c.winfo_height(), 330)
        left, top, bottom = 185, 28, h - 28
        maxv = max([v for _, v in items] + [1])
        usable = max(w - left - 55, 100)
        row_h = max((bottom - top) / max(len(items), 1), 42)

        for i, (name, value) in enumerate(items):
            y = top + i * row_h + 8
            c.create_text(10, y + 11, text=name, anchor="w",
                          font=("Segoe UI", 10), fill="#33485C")
            bar_w = (value / maxv) * usable
            c.create_rectangle(left, y, left + bar_w, y + 23,
                               fill="#2C6E9E", outline="")
            c.create_text(left + bar_w + 8, y + 11, text=str(value),
                          anchor="w", font=("Segoe UI", 10, "bold"), fill="#17324D")

    def refresh_keys(self):
        query=self.key_search.get().strip().lower() if hasattr(self,"key_search") else ""
        for row in self.key_tree.get_children(): self.key_tree.delete(row)
        for item in self.repo.list_keys():
            blob=" ".join(str(item.get(k,"") or "") for k in ["Número no Claviculário","Código","Descrição","Categoria/Módulo","Local","Cópias"]).lower()
            if query and query not in blob: continue
            numero=item.get("Número no Claviculário"); numero_txt=f"{int(numero):03d}" if numero not in (None,"") else ""
            stock=self.repo.key_stock_info(item.get("Código")) or {"copias":0,"disponiveis":1,"status":"Disponível"}
            copies_txt=str(stock["copias"]) if stock["copias"]>0 else "Sem cópia"
            self.key_tree.insert("","end",values=(numero_txt,item.get("Código",""),item.get("Descrição",""),item.get("Categoria/Módulo",""),item.get("Local",""),copies_txt,stock["disponiveis"],stock["status"]))

    def refresh_withdraw_combo(self):
        available=[]
        for k in self.repo.list_keys():
            if k.get("Número no Claviculário") in (None,""): continue
            stock=self.repo.key_stock_info(k.get("Código"))
            if stock and stock["disponiveis"]>0:
                available.append(f"{int(k.get('Número no Claviculário')):03d} | {k.get('Código')}")
        self.withdraw_combo["values"]=available
        current=self.withdraw_vars["codigo"].get()
        if current and current not in available:
            self.withdraw_vars["codigo"].set(""); self.withdraw_info.set("Selecione uma Chave do ECL com unidade disponível.")

    def refresh_returns(self):
        for row in self.return_tree.get_children():
            self.return_tree.delete(row)
        for item in self.repo.list_open_withdrawals():
            numero = item.get("Número no Claviculário")
            numero_txt = f"{int(numero):03d}" if numero not in (None, "") else ""
            self.return_tree.insert("", "end", values=(
                item.get("Protocolo", ""),
                numero_txt,
                item.get("Código ECL", ""),
                item.get("Descrição", ""),
                item.get("Responsável", ""),
                item.get("Setor", ""),
                item.get("Data/Hora Retirada", ""),
            ))

    def refresh_protocols(self):
        query = self.protocol_search.get() if hasattr(self, "protocol_search") else ""
        for row in self.protocol_tree.get_children():
            self.protocol_tree.delete(row)
        for item in self.repo.list_protocols(query):
            numero = item.get("Número no Claviculário")
            numero_txt = f"{int(numero):03d}" if numero not in (None, "") else ""
            self.protocol_tree.insert("", "end", values=(
                item.get("Protocolo", ""),
                numero_txt,
                item.get("Código ECL", ""),
                item.get("Descrição", ""),
                item.get("Responsável", ""),
                item.get("Data/Hora Retirada", ""),
                item.get("Status", ""),
            ))

    def refresh_history(self):
        query = self.history_search.get() if hasattr(self, "history_search") else ""
        for row in self.history_tree.get_children():
            self.history_tree.delete(row)
        for item in self.repo.list_movements(query):
            numero = item.get("Número no Claviculário")
            numero_txt = f"{int(numero):03d}" if numero not in (None, "") else ""
            self.history_tree.insert("", "end", values=(
                item.get("Protocolo", ""),
                numero_txt,
                item.get("Código ECL", ""),
                item.get("Responsável", ""),
                item.get("Data/Hora Retirada", ""),
                item.get("Data/Hora Devolução", ""),
                item.get("Status", ""),
            ))

    def refresh_inactive(self):
        for row in self.inactive_tree.get_children():
            self.inactive_tree.delete(row)
        for item in self.repo.list_keys(include_inactive=True):
            if str(item.get("Ativa", "SIM")).upper() == "SIM":
                continue
            numero = item.get("Número no Claviculário")
            numero_txt = f"{int(numero):03d}" if numero not in (None, "") else ""
            self.inactive_tree.insert("", "end", values=(
                numero_txt,
                item.get("Código", ""),
                item.get("Descrição", ""),
                item.get("Categoria/Módulo", ""),
                item.get("Local", ""),
                item.get("Motivo inativação", ""),
            ))


def main():
    try:
        app = App()
        app.mainloop()
    except Exception as exc:
        log_error(exc)
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Erro inesperado", str(exc))
            root.destroy()
        except Exception:
            print(str(exc), file=sys.stderr)


if __name__ == "__main__":
    main()
